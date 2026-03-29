"""
Multi-format document parser.
Supports: .txt, .pdf, .docx, .json, .xlsx, .xls, .csv
"""
import json
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def parse_document(file_path: str, file_type: str) -> str:
    """Route to the correct parser based on file extension."""
    parsers = {
        ".txt": _parse_txt,
        ".pdf": _parse_pdf,
        ".docx": _parse_docx,
        ".json": _parse_json,
        ".xlsx": _parse_xlsx,
        ".xls": _parse_xlsx,
        ".csv": _parse_csv,
    }
    ext = file_type if file_type.startswith(".") else f".{file_type}"
    parser = parsers.get(ext.lower())
    if not parser:
        raise ValueError(f"Unsupported file type: {ext}")
    return parser(file_path)


def _parse_txt(path: str) -> str:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    except UnicodeDecodeError:
        with open(path, "r", encoding="latin-1") as f:
            return f.read()


def _parse_pdf(path: str) -> str:
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(path)
        texts = []
        for page in doc:
            texts.append(page.get_text())
        return "\n".join(texts)
    except Exception as e:
        logger.error(f"PyMuPDF failed: {e}, trying pdfplumber")
        try:
            import pdfplumber
            with pdfplumber.open(path) as pdf:
                return "\n".join(p.extract_text() or "" for p in pdf.pages)
        except Exception as e2:
            raise RuntimeError(f"PDF parsing failed: {e2}")


def _parse_docx(path: str) -> str:
    from docx import Document
    doc = Document(path)
    paragraphs = []
    for para in doc.paragraphs:
        if para.text.strip():
            # Preserve heading hierarchy
            if para.style.name.startswith("Heading"):
                paragraphs.append(f"\n## {para.text.strip()}\n")
            else:
                paragraphs.append(para.text.strip())
    return "\n".join(paragraphs)


def _parse_json(path: str) -> str:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return _flatten_json(data)


def _flatten_json(obj, prefix="") -> str:
    """Recursively flatten nested JSON into key: value lines."""
    lines = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            key = f"{prefix}.{k}" if prefix else k
            lines.append(_flatten_json(v, key))
    elif isinstance(obj, list):
        for i, item in enumerate(obj):
            key = f"{prefix}[{i}]"
            lines.append(_flatten_json(item, key))
    else:
        lines.append(f"{prefix}: {obj}")
    return "\n".join(lines)


def _parse_xlsx(path: str) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_text = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_text.append(f"=== Sheet: {sheet_name} ===")
        headers = None
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if all(cell is None for cell in row):
                continue
            if row_idx == 0:
                headers = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(row)]
                all_text.append(" | ".join(headers))
            else:
                if headers:
                    row_text = " | ".join(
                        f"{headers[i]}: {str(v) if v is not None else ''}"
                        for i, v in enumerate(row) if i < len(headers)
                    )
                else:
                    row_text = " | ".join(str(v) if v is not None else "" for v in row)
                all_text.append(row_text)
    return "\n".join(all_text)


def _parse_csv(path: str) -> str:
    import pandas as pd
    try:
        df = pd.read_csv(path)
    except Exception:
        df = pd.read_csv(path, encoding="latin-1", sep=None, engine="python")
    return df.to_string(index=False)
